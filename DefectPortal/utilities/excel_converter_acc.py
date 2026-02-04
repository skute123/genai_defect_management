import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os

# -----------------------------
# Paths
# -----------------------------
base_dir = os.path.dirname(os.path.abspath(__file__))

input_folder = os.path.join(base_dir, "../sheet")
output_folder = os.path.join(base_dir, "../output")
os.makedirs(output_folder, exist_ok=True)

# Input files
first_csv_file = os.path.join(input_folder, 'defect_sheet_acc.csv')
second_xlsx_file = os.path.join(input_folder, 'ttwos_extract_acc.xlsx')

# Output file
output_excel_file = os.path.join(output_folder, 'filtered_output_acc.xlsx')

# -----------------------------
# Settings
# -----------------------------
columns_to_extract = [
    'Summary', 'Issue key', 'Priority', 'Resolution', 'Fix Version/s',
    'Description', 'Custom field (OSF-Fix Description)', 'Custom field (OSF-Stack)',
    'Custom field (OSF-System)', 'Custom field (Vendor + Application)', 'Comment'
]

second_sheet_mapping = {
    'Ticketnummer': 'Issue key',
    'Prio': 'Priority',
    'Buchungsdatum': 'Start Date',
    'Kurzbeschreibung': 'Summary',
    'Beschreibung': 'Description',
    'Rückmeldebeschreibung': 'Comment',
    'Kategorie1 +': 'Custom field (OSF-System)',
    'Kategorie2 +': 'Custom field (OSF-Stack)',
    'Kategorie3 +': 'Custom field (Vendor + Application)'
}

# -----------------------------
# Helper: process Jira-like CSV
# -----------------------------
def process_csv(file, columns_to_extract):
    df = pd.read_csv(file)

    # Merge duplicate columns (.1, .2 suffixes)
    grouped_cols = {}
    for col in df.columns:
        base = re.sub(r'\.\d+$', '', col)
        grouped_cols.setdefault(base, []).append(col)

    merged_columns = {}
    for base_col, cols in grouped_cols.items():
        if len(cols) == 1:
            merged_columns[base_col] = df[cols[0]]
        else:
            merged_columns[base_col] = df[cols].apply(
                lambda row: '\n '.join([str(val).strip() for val in row if pd.notna(val) and str(val).strip()]),
                axis=1
            )

    merged_df = pd.DataFrame(merged_columns)
    return merged_df.reindex(columns=columns_to_extract, fill_value="")

# -----------------------------
# Main Conversion Logic
# -----------------------------
df1_aligned = process_csv(first_csv_file, columns_to_extract)

if os.path.exists(second_xlsx_file):
    print("TTWOS extract found. Combining with Jira defects...")
    df2 = pd.read_excel(second_xlsx_file)
    df2 = df2.rename(columns=second_sheet_mapping)
    df2_aligned = df2.reindex(columns=columns_to_extract, fill_value="")
    combined_df = pd.concat([df1_aligned, df2_aligned], ignore_index=True)
else:
    print("TTWOS extract not found. Proceeding with Jira CSV only...")
    combined_df = df1_aligned

combined_df.to_excel(output_excel_file, index=False)

# -----------------------------
# Formatting
# -----------------------------
def apply_formatting(file):
    wb = load_workbook(file)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    fixed_width = 30
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = fixed_width

    wb.save(file)

apply_formatting(output_excel_file)

print(f" ACC Excel file saved as: {output_excel_file}")
