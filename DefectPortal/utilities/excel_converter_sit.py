import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os

# -----------------------------
# Paths
# -----------------------------
base_dir = os.path.dirname(os.path.abspath(__file__))

input_folder = os.path.join(base_dir, "../sheet")
output_folder = os.path.join(base_dir, "../output")
os.makedirs(output_folder, exist_ok=True)

# Input file
third_csv_file = os.path.join(input_folder, 'defect_sheet_sit.csv')

# Output file
output_excel_file_sit = os.path.join(output_folder, 'filtered_output_sit.xlsx')

# -----------------------------
# Settings
# -----------------------------
columns_to_extract = [
    'Summary', 'Issue key', 'Priority', 'Resolution', 'Fix Version/s',
    'Description', 'Custom field (OSF-Fix Description)', 'Custom field (OSF-Stack)',
    'Custom field (OSF-System)', 'Custom field (Vendor + Application)', 'Comment'
]

# -----------------------------
# Helper: process Jira-like CSV
# -----------------------------
def process_csv(file, columns_to_extract):
    df = pd.read_csv(file)

    # Merge duplicate columns (.1, .2 suffixes)
    grouped_cols = {}
    for col in df.columns:
        base = re.sub(r'\.\d+$', '', col)
        grouped_cols.setdefault(base, []).append(col)

    merged_columns = {}
    for base_col, cols in grouped_cols.items():
        if len(cols) == 1:
            merged_columns[base_col] = df[cols[0]]
        else:
            merged_columns[base_col] = df[cols].apply(
                lambda row: '\n '.join([str(val).strip() for val in row if pd.notna(val) and str(val).strip()]),
                axis=1
            )

    merged_df = pd.DataFrame(merged_columns)
    return merged_df.reindex(columns=columns_to_extract, fill_value="")

# -----------------------------
# Main Conversion Logic
# -----------------------------
df3_aligned = process_csv(third_csv_file, columns_to_extract)
df3_aligned.to_excel(output_excel_file_sit, index=False)

# -----------------------------
# Formatting
# -----------------------------
def apply_formatting(file):
    wb = load_workbook(file)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    fixed_width = 30
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = fixed_width

    wb.save(file)

apply_formatting(output_excel_file_sit)

print(f"✅ SIT Excel file saved as: {output_excel_file_sit}")
