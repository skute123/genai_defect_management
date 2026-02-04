# columns_to_extract = ['Summary', 'Issue key', 'Priority', 'Resolution','Fix Version/s', 'Description','Custom field (OSF-Fix Description)','Custom field (OSF-Stack)','Custom field (OSF-System)','Custom field (Vendor + Application)','Comment']

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Step 1: Read CSV
input_csv_file = 'defect_sheet.csv'
df = pd.read_csv(input_csv_file)

# Step 2: Group columns by base name (remove .1, .2 suffixes)
# comment.1, comment.2, to remove . we have done this 
grouped_cols = {}
for col in df.columns:
    base = re.sub(r'\.\d+$', '', col)  # Remove .1, .2, etc.
    grouped_cols.setdefault(base, []).append(col)

# Step 3: Merge values for each group,  merge the values of same name col into a single column with new line \n
merged_columns = {}
for base_col, cols in grouped_cols.items():
    if len(cols) == 1:
        merged_columns[base_col] = df[cols[0]]
    else:
        merged_columns[base_col] = df[cols].apply(
            lambda row: '\n '.join(
                [str(val).strip() for val in row if pd.notna(val) and str(val).strip()]
            ),
            axis=1
        )

# Step 4: Convert to DataFrame
merged_df = pd.DataFrame(merged_columns)

# Step 5: Select specific columns
columns_to_extract = ['Summary', 'Issue key', 'Priority', 'Resolution','Fix Version/s', 'Description','Custom field (OSF-Fix Description)','Custom field (OSF-Stack)','Custom field (OSF-System)','Custom field (Vendor + Application)','Comment']
filtered_df = merged_df[columns_to_extract]

# Step 6: Save to Excel
output_excel_file = 'filtered_output.xlsx'
filtered_df.to_excel(output_excel_file, index=False)

# Step 7: Apply formatting
wb = load_workbook(output_excel_file)
ws = wb.active

# Word wrap and top alignment
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Fixed column width
fixed_width = 30
for col in ws.columns:
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = fixed_width

wb.save(output_excel_file)

print(f" Cleaned and formatted Excel saved as: {output_excel_file}")